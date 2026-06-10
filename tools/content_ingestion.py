"""Content Ingestion — turn raw exam PDFs/spreadsheets into Exam Packs.

Uses QwenPaw's PDF and XLSX skill patterns:
  - pypdf / pdfplumber for past-question PDFs
  - openpyxl for syllabus spreadsheets
  - Outputs: pack.json files for the exam_packs/ directory

Usage:
    from stodi.tools.content_ingestion import (
        ingest_past_questions_pdf,
        ingest_syllabus_spreadsheet,
        build_exam_pack,
    )
"""

import json
import re
from pathlib import Path
from datetime import datetime

from stodi.config.exam_pack import (
    ExamPack,
    Syllabus,
    Topic,
    PastQuestion,
    PastQuestionCorpus,
    RetentionCadence,
)


# ═══════════════════════════════════════════════════════════════
# PDF PAST-QUESTION INGESTION
# ═══════════════════════════════════════════════════════════════

def ingest_past_questions_pdf(
    pdf_path: str,
    exam_board: str,
    subject: str,
    year: int,
    topic_map: dict[str, str] | None = None,
) -> dict:
    """Extract past questions from a WAEC/JAMB exam paper PDF.

    Args:
        pdf_path: Path to the PDF file
        exam_board: "WAEC", "JAMB", etc.
        subject: "Mathematics", "English Language", etc.
        year: Exam year
        topic_map: Optional mapping of question numbers → topic IDs
                   e.g. {"1": "A4", "2": "B3", ...}

    Returns:
        Summary of extracted questions + list of PastQuestion objects
    """
    path = Path(pdf_path).expanduser()
    if not path.exists():
        return {"error": f"File not found: {pdf_path}"}

    try:
        import pdfplumber
    except ImportError:
        return {"error": "pdfplumber not installed: pip install pdfplumber"}

    questions = []
    q_counter = 0

    with pdfplumber.open(str(path)) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"

    # Parse questions from the extracted text
    # Common WAEC format: "1. " or "(a)" or "Question 1"
    raw_questions = _split_into_questions(full_text)

    for i, q_text in enumerate(raw_questions):
        if not q_text.strip():
            continue

        q_counter += 1
        q_id = f"{exam_board}-{subject.upper()[:4]}-{year}-{q_counter:02d}"

        # Detect question type
        q_type = _detect_question_type(q_text)

        # Detect if MCQ (has options A-D)
        options = _extract_mcq_options(q_text)
        correct = None

        # Map to topic if map provided
        topic_id = "UNKNOWN"
        if topic_map and str(i + 1) in topic_map:
            topic_id = topic_map[str(i + 1)]

        pq = PastQuestion(
            id=q_id,
            year=year,
            subject=subject,
            topic_id=topic_id,
            question_type=q_type,
            question_text=q_text[:500].strip(),
            options=options,
            correct_answer=correct,
            marking_scheme=None,
            marks=1 if q_type == "mcq" else 3,
        )
        questions.append(pq.model_dump())

    return {
        "source": str(path),
        "exam_board": exam_board,
        "subject": subject,
        "year": year,
        "questions_extracted": len(questions),
        "questions": questions,
        "raw_text_length": len(full_text),
    }


def _split_into_questions(text: str) -> list[str]:
    """Split exam paper text into individual questions."""
    # Pattern: "1." or "2." at start of line, or "(a)" subsections
    parts = re.split(r'\n(?=\d{1,2}\.\s)', text)
    if len(parts) <= 1:
        # Fallback: split by double newlines
        parts = [p.strip() for p in text.split('\n\n') if p.strip()]
    return [p.strip() for p in parts if p.strip()]


def _detect_question_type(text: str) -> str:
    """Detect whether a question is MCQ, theory, or essay."""
    has_options = bool(re.search(r'\b[A-D]\.\s', text))
    if has_options:
        return "mcq"
    if len(text) > 300:
        return "essay"
    return "theory"


def _extract_mcq_options(text: str) -> list[str] | None:
    """Extract A/B/C/D options from MCQ text."""
    options = re.findall(r'([A-D])\.\s*(.*?)(?=\s*[A-D]\.|$)', text)
    if options and len(options) >= 2:
        return [f"{letter}. {text.strip()}" for letter, text in options]
    return None


# ═══════════════════════════════════════════════════════════════
# SYLLABUS SPREADSHEET INGESTION
# ═══════════════════════════════════════════════════════════════

def ingest_syllabus_spreadsheet(
    file_path: str,
    exam_board: str,
    subject: str,
    name_col: str = "topic",
    id_col: str = "id",
    difficulty_col: str = "difficulty",
    weight_col: str = "weight",
    parent_col: str | None = "parent_id",
) -> dict:
    """Parse a syllabus spreadsheet into a topic tree.

    Accepts .xlsx, .csv, .tsv files.

    Args:
        file_path: Path to the spreadsheet
        exam_board: "WAEC", "JAMB", etc.
        subject: Subject name
        name_col: Column name for topic names
        id_col: Column name for topic IDs
        difficulty_col: Column name for difficulty
        weight_col: Column name for weight/marks allocation
        parent_col: Column name for parent topic ID (optional)
    """
    path = Path(file_path).expanduser()
    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    suffix = path.suffix.lower()

    if suffix == '.csv':
        import csv
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    elif suffix in ('.xlsx', '.xlsm'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            wb.close()
        except ImportError:
            return {"error": "openpyxl not installed: pip install openpyxl"}
    elif suffix == '.tsv':
        import csv
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter='\t')
            rows = list(reader)
    else:
        return {"error": f"Unsupported format: {suffix}. Use .xlsx, .csv, or .tsv"}

    topics = []
    for row in rows:
        topic = Topic(
            id=str(row.get(id_col, f"T{len(topics)+1}")).strip(),
            name=str(row.get(name_col, "Unknown")).strip(),
            parent_id=str(row.get(parent_col, "")).strip() or None if parent_col else None,
            difficulty=str(row.get(difficulty_col, "medium")).strip().lower(),
            weight=float(row.get(weight_col, 1.0)),
        )
        topics.append(topic.model_dump())

    return {
        "source": str(path),
        "exam_board": exam_board,
        "subject": subject,
        "topics_extracted": len(topics),
        "topics": topics,
    }


# ═══════════════════════════════════════════════════════════════
# BUILD COMPLETE EXAM PACK FROM INGESTED DATA
# ═══════════════════════════════════════════════════════════════

def build_exam_pack(
    exam_board: str,
    subject: str,
    topics: list[dict],
    questions: list[dict],
    output_path: str | None = None,
    retention_cadence: dict | None = None,
) -> dict:
    """Assemble a complete Exam Pack from ingested topics + questions.

    Args:
        exam_board: "WAEC", "JAMB", etc.
        subject: Subject name
        topics: List of topic dicts (from ingest_syllabus_spreadsheet)
        questions: List of question dicts (from ingest_past_questions_pdf)
        output_path: Where to save pack.json (None = auto)
        retention_cadence: Optional custom retention config
    """
    # Build syllabus
    syllabus = Syllabus(
        subject=subject,
        exam_board=exam_board,
        topics=[Topic(**t) for t in topics],
        total_marks=100,
    )

    # Build question corpus
    corpus = PastQuestionCorpus(
        subject=subject,
        exam_board=exam_board,
        questions=[PastQuestion(**q) for q in questions],
    )

    # Build retention config
    cadence = RetentionCadence(**(retention_cadence or {}))

    # Assemble pack
    pack = ExamPack(
        name=f"{exam_board} {subject}",
        exam_board=exam_board,
        subject=subject,
        version="1.0",
        syllabus=syllabus,
        past_questions=corpus,
        retention=cadence,
    )

    # Determine output path
    if output_path is None:
        from stodi.config.exam_pack import PACKS_DIR
        output_path = str(
            PACKS_DIR / exam_board.lower() / subject.lower() / "pack.json"
        )

    # Save
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    pack_dict = pack.model_dump()
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(pack_dict, f, indent=2, ensure_ascii=False)

    return {
        "status": "built",
        "pack_name": pack.name,
        "topics": len(syllabus.topics),
        "questions": len(corpus.questions),
        "output": str(out),
    }
